from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Header, Query
from fastapi.responses import Response, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import requests
from emergentintegrations.llm.chat import LlmChat, UserMessage
import base64
import io
from PIL import Image
import csv
import json
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Object Storage Configuration
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "kairos-accounting"
storage_key = None

# Initialize storage
def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str) -> tuple:
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# Create the main app
app = FastAPI(title="Kairos Accounting")

# Create API router
api_router = APIRouter(prefix="/api")

# Pydantic Models
class ChartOfAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ledger_name: str
    category: str
    opening_balance: float = 0.0
    current_balance: float = 0.0
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CostCenter(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TransactionBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_prompt: Optional[str] = None
    user_id: str = "default_user"
    module: str
    status: str = "draft"
    posting_date: Optional[str] = None
    service_period_start: Optional[str] = None
    service_period_end: Optional[str] = None
    business_unit: Optional[str] = None
    cost_center: str = "General"
    extracted_data: Optional[Dict[str, Any]] = None
    journal_entries: Optional[List[Dict[str, Any]]] = None
    document_id: Optional[str] = None
    import_source: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    posted_at: Optional[str] = None

class DocumentUpload(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    storage_path: str
    original_filename: str
    content_type: str
    size: int
    extracted_data: Optional[Dict[str, Any]] = None
    transaction_id: Optional[str] = None
    is_deleted: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class JournalEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_id: str
    account: str
    debit: float = 0.0
    credit: float = 0.0
    description: str
    posting_date: str
    cost_center: str = "General"
    tax_details: Optional[Dict[str, Any]] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class VendorClient(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str
    name: str
    gstin: Optional[str] = None
    pan: Optional[str] = None
    legal_name: Optional[str] = None
    constitution: Optional[str] = None
    status: str = "Active"
    address: Optional[str] = None
    contact: Optional[str] = None
    email: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_name: str
    hsn_code: Optional[str] = None
    unit: str = "pcs"
    current_stock: float = 0.0
    landed_cost: float = 0.0
    valuation_method: str = "FIFO"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PromptRequest(BaseModel):
    prompt: str
    module: str
    user_id: str = "default_user"
    document_id: Optional[str] = None
    cost_center: str = "General"

class PostTransactionRequest(BaseModel):
    transaction_id: str
    cost_center: Optional[str] = None

class ReportQuery(BaseModel):
    query: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class CSVImportRequest(BaseModel):
    module: str
    csv_data: str

# GSTIN Lookup (Mock implementation - replace with real API)
async def lookup_gstin(gstin: str) -> Dict[str, Any]:
    """Mock GSTIN lookup - in production, integrate with government API"""
    # This would call actual GSTIN verification API
    return {
        "legal_name": f"Mock Company for {gstin}",
        "constitution": "Private Limited",
        "status": "Active",
        "address": "Mock Address, India"
    }

# AI Services
async def parse_prompt_with_ai(prompt: str, module: str, extracted_ocr: Optional[Dict] = None, cost_center: str = "General") -> Dict[str, Any]:
    """Use Claude Sonnet 4.5 to parse user prompt and generate journal entries"""
    try:
        chat = LlmChat(
            api_key=EMERGENT_KEY,
            session_id=f"prompt-{uuid.uuid4()}",
            system_message="""You are an expert Indian ERP accountant for Kairos Accounting system. Parse the user's natural language prompt and extract:
            1. posting_date (format: YYYY-MM-DD)
            2. service_period_start (format: YYYY-MM-DD)
            3. service_period_end (format: YYYY-MM-DD)
            4. business_unit (e.g., Mumbai Office, Gujarat Plant)
            5. journal_entries: List of structured entries with:
               - account: Ledger name
               - debit: Amount (use 0 if credit entry)
               - credit: Amount (use 0 if debit entry)
               - description: Brief description
               - tax_details: {igst: 0, cgst: 0, sgst: 0, cess: 0, tds: 0}
            
            For GST transactions:
            - Determine if interstate (IGST) or intrastate (CGST + SGST)
            - Default GST rate: 18%
            - Include proper GST accounts: Input GST / Output GST
            - Calculate landed cost for goods = Purchase Price + IGST/CGST+SGST + Freight
            
            For TDS:
            - Identify salary payments and apply appropriate TDS
            - Include TDS Payable account
            
            Return ONLY a valid JSON object with these fields. No markdown, no explanations."""
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")
        
        prompt_text = f"Module: {module}\nCost Center: {cost_center}\nPrompt: {prompt}"
        if extracted_ocr:
            prompt_text += f"\n\nExtracted OCR Data: {json.dumps(extracted_ocr)}"
        
        user_message = UserMessage(text=prompt_text)
        response = await chat.send_message(user_message)
        
        result = json.loads(response)
        return result
    except Exception as e:
        logging.error(f"AI parsing error: {e}")
        return {
            "posting_date": datetime.now(timezone.utc).date().isoformat(),
            "service_period_start": datetime.now(timezone.utc).date().isoformat(),
            "service_period_end": datetime.now(timezone.utc).date().isoformat(),
            "business_unit": "Main Office",
            "journal_entries": []
        }

async def extract_document_data(image_data: bytes, filename: str) -> Dict[str, Any]:
    """Use Gemini 3 vision for OCR extraction"""
    try:
        img = Image.open(io.BytesIO(image_data))
        buffered = io.BytesIO()
        img.save(buffered, format="PNG")
        img_base64 = base64.b64encode(buffered.getvalue()).decode()
        
        chat = LlmChat(
            api_key=EMERGENT_KEY,
            session_id=f"ocr-{uuid.uuid4()}",
            system_message="""You are an OCR expert for Indian financial documents. Extract:
            - Vendor/Client name
            - GSTIN (15-character GST identification number)
            - Invoice number
            - Invoice date (format: YYYY-MM-DD)
            - Total amount
            - Line items with: item_name, hsn_code, quantity, rate, amount
            - GST breakdown: {cgst_rate, cgst_amount, sgst_rate, sgst_amount, igst_rate, igst_amount, cess}
            - TDS if applicable
            - Freight charges if any
            
            Return ONLY valid JSON. No markdown."""
        ).with_model("gemini", "gemini-3-flash-preview")
        
        user_message = UserMessage(text=f"Extract all data from this invoice/receipt image. Return JSON only.")
        response = await chat.send_message(user_message)
        
        result = json.loads(response)
        return result
    except Exception as e:
        logging.error(f"OCR extraction error: {e}")
        return {
            "vendor_name": "Unknown",
            "gstin": "",
            "invoice_number": "",
            "invoice_date": "",
            "total_amount": 0,
            "line_items": [],
            "gst_breakdown": {}
        }

# CSV Validation and Processing
async def validate_csv_data(csv_data: str, module: str) -> Dict[str, Any]:
    """Validate CSV data against Zoho-standard headers"""
    required_headers = {
        "purchases": ["Date", "Entity Name", "Item/Service", "Rate", "GST Rate", "Total"],
        "sales": ["Date", "Entity Name", "Item/Service", "Rate", "GST Rate", "Total"],
        "journals": ["Date", "Ledger", "Debit", "Credit", "Description"],
        "payments": ["Date", "Entity Name", "Amount", "Payment Mode"]
    }
    
    errors = []
    warnings = []
    
    try:
        csv_file = StringIO(csv_data)
        reader = csv.DictReader(csv_file)
        headers = reader.fieldnames
        
        # Check required headers
        module_key = module.replace("-", "")
        if module_key in required_headers:
            for req_header in required_headers[module_key]:
                if req_header not in headers:
                    errors.append(f"Missing required header: {req_header}")
        
        # Validate data rows
        rows = list(reader)
        for idx, row in enumerate(rows, start=2):
            # Validate date format
            if "Date" in row and row["Date"]:
                try:
                    datetime.strptime(row["Date"], "%Y-%m-%d")
                except:
                    errors.append(f"Row {idx}: Invalid date format. Use YYYY-MM-DD")
            
            # Validate ledger existence for journals
            if module == "journals" and "Ledger" in row:
                ledger_exists = await db.chart_of_accounts.find_one({"ledger_name": row["Ledger"]}, {"_id": 0})
                if not ledger_exists:
                    warnings.append(f"Row {idx}: Ledger '{row['Ledger']}' not found in Chart of Accounts")
        
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "row_count": len(rows)
        }
    except Exception as e:
        return {
            "valid": False,
            "errors": [f"CSV parsing error: {str(e)}"],
            "warnings": [],
            "row_count": 0
        }

# API Endpoints
@api_router.get("/")
async def root():
    return {"message": "Kairos Accounting API", "version": "2.0.0"}

# Chart of Accounts
@api_router.post("/coa/upload")
async def upload_coa(file: UploadFile = File(...)):
    """Bulk upload Chart of Accounts from CSV"""
    try:
        content = await file.read()
        csv_data = content.decode('utf-8')
        csv_file = StringIO(csv_data)
        reader = csv.DictReader(csv_file)
        
        imported = 0
        errors = []
        
        for idx, row in enumerate(reader, start=2):
            try:
                coa = ChartOfAccount(
                    ledger_name=row.get("Ledger Name", ""),
                    category=row.get("Category", "Other"),
                    opening_balance=float(row.get("Opening Balance", 0))
                )
                coa_dict = coa.model_dump()
                coa_dict["current_balance"] = coa_dict["opening_balance"]
                await db.chart_of_accounts.insert_one(coa_dict)
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {"imported": imported, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/coa")
async def get_coa():
    """Get all Chart of Accounts"""
    coa_list = await db.chart_of_accounts.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return coa_list

@api_router.post("/coa")
async def create_coa(coa: ChartOfAccount):
    """Create single Chart of Account entry"""
    coa_dict = coa.model_dump()
    coa_dict["current_balance"] = coa_dict["opening_balance"]
    await db.chart_of_accounts.insert_one(coa_dict)
    return coa_dict

# Cost Centers
@api_router.post("/cost-centers")
async def create_cost_center(cc: CostCenter):
    """Create cost center"""
    await db.cost_centers.insert_one(cc.model_dump())
    return cc.model_dump()

@api_router.get("/cost-centers")
async def get_cost_centers():
    """Get all cost centers"""
    centers = await db.cost_centers.find({"is_active": True}, {"_id": 0}).to_list(100)
    return centers

# Vendor/Client Master
@api_router.post("/entities")
async def create_entity(entity: VendorClient):
    """Create vendor or client with optional GSTIN lookup"""
    entity_dict = entity.model_dump()
    
    # Auto-lookup GSTIN if provided
    if entity.gstin and not entity.legal_name:
        gstin_data = await lookup_gstin(entity.gstin)
        entity_dict.update(gstin_data)
    
    await db.entities.insert_one(entity_dict)
    return entity_dict

@api_router.get("/entities")
async def get_entities(entity_type: Optional[str] = None):
    """Get vendors or clients"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    entities = await db.entities.find(query, {"_id": 0}).to_list(1000)
    return entities

# CSV Import
@api_router.post("/import/validate")
async def validate_import(request: CSVImportRequest):
    """Validate CSV before import"""
    validation = await validate_csv_data(request.csv_data, request.module)
    return validation

@api_router.post("/import/process")
async def process_import(file: UploadFile = File(...), module: str = Query(...)):
    """Process CSV import for any module"""
    try:
        content = await file.read()
        csv_data = content.decode('utf-8')
        
        # Validate first
        validation = await validate_csv_data(csv_data, module)
        if not validation["valid"]:
            return {"success": False, "validation": validation}
        
        # Process import
        csv_file = StringIO(csv_data)
        reader = csv.DictReader(csv_file)
        imported = 0
        
        for row in reader:
            # Create transaction based on module
            txn = TransactionBase(
                module=module,
                posting_date=row.get("Date"),
                import_source="csv",
                status="draft"
            )
            
            # Build journal entries from CSV
            entries = []
            if module == "journals":
                entries.append({
                    "account": row.get("Ledger"),
                    "debit": float(row.get("Debit", 0)),
                    "credit": float(row.get("Credit", 0)),
                    "description": row.get("Description", "")
                })
            elif module in ["purchases", "sales"]:
                # Auto-generate journal entries from purchase/sale data
                total = float(row.get("Total", 0))
                gst_rate = float(row.get("GST Rate", 0))
                base = total / (1 + gst_rate / 100)
                gst_amt = total - base
                
                if module == "purchases":
                    entries.append({"account": "Purchases", "debit": base, "credit": 0, "description": row.get("Item/Service")})
                    entries.append({"account": "Input GST", "debit": gst_amt, "credit": 0, "description": "GST"})
                    entries.append({"account": row.get("Entity Name"), "debit": 0, "credit": total, "description": "Vendor"})
                else:
                    entries.append({"account": row.get("Entity Name"), "debit": total, "credit": 0, "description": "Customer"})
                    entries.append({"account": "Sales", "debit": 0, "credit": base, "description": row.get("Item/Service")})
                    entries.append({"account": "Output GST", "debit": 0, "credit": gst_amt, "description": "GST"})
                
                # Update inventory if Item contains HSN
                if row.get("HSN Code"):
                    item = await db.inventory.find_one({"hsn_code": row.get("HSN Code")}, {"_id": 0})
                    if not item:
                        # Create new inventory item
                        new_item = InventoryItem(
                            item_name=row.get("Item/Service"),
                            hsn_code=row.get("HSN Code"),
                            landed_cost=base,
                            current_stock=float(row.get("Quantity", 1))
                        )
                        await db.inventory.insert_one(new_item.model_dump())
                    else:
                        # Update stock
                        qty = float(row.get("Quantity", 1))
                        await db.inventory.update_one(
                            {"hsn_code": row.get("HSN Code")},
                            {"$inc": {"current_stock": qty}}
                        )
            
            txn_dict = txn.model_dump()
            txn_dict["journal_entries"] = entries
            await db.transactions.insert_one(txn_dict)
            imported += 1
        
        return {"success": True, "imported": imported, "validation": validation}
    except Exception as e:
        logging.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Document Upload
@api_router.post("/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    """Upload invoice/receipt and perform OCR"""
    try:
        ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
        doc_id = str(uuid.uuid4())
        path = f"{APP_NAME}/documents/{doc_id}.{ext}"
        
        data = await file.read()
        result = put_object(path, data, file.content_type or "application/octet-stream")
        
        extracted_data = await extract_document_data(data, file.filename)
        
        doc_record = DocumentUpload(
            id=doc_id,
            storage_path=result["path"],
            original_filename=file.filename,
            content_type=file.content_type,
            size=result["size"],
            extracted_data=extracted_data
        )
        
        await db.documents.insert_one(doc_record.model_dump())
        
        return {"document_id": doc_id, "extracted_data": extracted_data}
    except Exception as e:
        logging.error(f"Document upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documents/{doc_id}/download")
async def download_document(doc_id: str):
    """Download document file"""
    try:
        record = await db.documents.find_one({"id": doc_id, "is_deleted": False}, {"_id": 0})
        if not record:
            raise HTTPException(status_code=404, detail="Document not found")
        
        data, content_type = get_object(record["storage_path"])
        return Response(content=data, media_type=record.get("content_type", content_type))
    except Exception as e:
        logging.error(f"Document download error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Process Prompt
@api_router.post("/transactions/prompt")
async def process_prompt(request: PromptRequest):
    """Process natural language prompt and create draft transaction"""
    try:
        extracted_ocr = None
        if request.document_id:
            doc = await db.documents.find_one({"id": request.document_id}, {"_id": 0})
            if doc:
                extracted_ocr = doc.get("extracted_data")
        
        parsed_data = await parse_prompt_with_ai(request.prompt, request.module, extracted_ocr, request.cost_center)
        
        transaction = TransactionBase(
            user_prompt=request.prompt,
            user_id=request.user_id,
            module=request.module,
            status="draft",
            posting_date=parsed_data.get("posting_date"),
            service_period_start=parsed_data.get("service_period_start"),
            service_period_end=parsed_data.get("service_period_end"),
            business_unit=parsed_data.get("business_unit"),
            cost_center=request.cost_center,
            extracted_data=parsed_data,
            journal_entries=parsed_data.get("journal_entries", []),
            document_id=request.document_id
        )
        
        await db.transactions.insert_one(transaction.model_dump())
        
        return transaction.model_dump()
    except Exception as e:
        logging.error(f"Prompt processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/transactions/drafts")
async def get_drafts(module: Optional[str] = None):
    """Get all draft transactions"""
    query = {"status": "draft"}
    if module:
        query["module"] = module
    
    drafts = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return drafts

@api_router.post("/transactions/post")
async def post_transaction(request: PostTransactionRequest):
    """Post a draft transaction to ledger"""
    try:
        transaction = await db.transactions.find_one({"id": request.transaction_id}, {"_id": 0})
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        if transaction["status"] != "draft":
            raise HTTPException(status_code=400, detail="Transaction already posted")
        
        cost_center = request.cost_center or transaction.get("cost_center", "General")
        
        await db.transactions.update_one(
            {"id": request.transaction_id},
            {"$set": {"status": "posted", "posted_at": datetime.now(timezone.utc).isoformat(), "cost_center": cost_center}}
        )
        
        journal_entries = transaction.get("journal_entries", [])
        for entry in journal_entries:
            je = JournalEntry(
                transaction_id=request.transaction_id,
                account=entry.get("account", ""),
                debit=entry.get("debit", 0.0),
                credit=entry.get("credit", 0.0),
                description=entry.get("description", ""),
                posting_date=transaction["posting_date"],
                cost_center=cost_center,
                tax_details=entry.get("tax_details")
            )
            await db.journal_entries.insert_one(je.model_dump())
            
            # Update CoA balances
            account_name = entry.get("account")
            debit_amt = entry.get("debit", 0.0)
            credit_amt = entry.get("credit", 0.0)
            net_change = debit_amt - credit_amt
            
            await db.chart_of_accounts.update_one(
                {"ledger_name": account_name},
                {"$inc": {"current_balance": net_change}},
                upsert=False
            )
        
        return {"message": "Transaction posted successfully", "transaction_id": request.transaction_id}
    except Exception as e:
        logging.error(f"Post transaction error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/transactions/posted")
async def get_posted_transactions(module: Optional[str] = None, limit: int = 50):
    """Get posted transactions"""
    query = {"status": "posted"}
    if module:
        query["module"] = module
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("posted_at", -1).to_list(limit)
    return transactions

# Conversational Reporting
@api_router.post("/reports/query")
async def conversational_report(request: ReportQuery):
    """AI-powered conversational reporting"""
    try:
        query = {"status": "posted"}
        if request.start_date and request.end_date:
            query["posting_date"] = {"$gte": request.start_date, "$lte": request.end_date}
        
        transactions = await db.transactions.find(query, {"_id": 0}).to_list(1000)
        journal_entries = await db.journal_entries.find({}, {"_id": 0}).to_list(1000)
        
        chat = LlmChat(
            api_key=EMERGENT_KEY,
            session_id=f"report-{uuid.uuid4()}",
            system_message=f"""You are an AI financial analyst for Kairos Accounting system (India ERP). 
            Answer the user's query based on the provided transaction data.
            
            Transactions: {json.dumps(transactions[:100])}
            Journal Entries: {json.dumps(journal_entries[:100])}
            
            Provide clear, concise answers with numbers and insights."""
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")
        
        user_message = UserMessage(text=request.query)
        response = await chat.send_message(user_message)
        
        return {"query": request.query, "answer": response, "data_points": len(transactions)}
    except Exception as e:
        logging.error(f"Report query error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Financial Statements
@api_router.get("/reports/balance-sheet")
async def get_balance_sheet(as_of_date: Optional[str] = None, cost_center: Optional[str] = None):
    """Generate Balance Sheet"""
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).date().isoformat()
    
    query = {"posting_date": {"$lte": as_of_date}}
    if cost_center:
        query["cost_center"] = cost_center
    
    entries = await db.journal_entries.find(query, {"_id": 0}).to_list(10000)
    
    balances = {}
    for entry in entries:
        account = entry["account"]
        if account not in balances:
            balances[account] = {"debit": 0, "credit": 0}
        balances[account]["debit"] += entry["debit"]
        balances[account]["credit"] += entry["credit"]
    
    assets = {k: v for k, v in balances.items() if "Asset" in k or "Bank" in k or "Cash" in k}
    liabilities = {k: v for k, v in balances.items() if "Liability" in k or "Payable" in k}
    equity = {k: v for k, v in balances.items() if "Equity" in k or "Capital" in k}
    
    # TB Integrity Check
    total_debit = sum(b["debit"] for b in balances.values())
    total_credit = sum(b["credit"] for b in balances.values())
    difference = abs(total_debit - total_credit)
    has_suspense = difference > 0.01
    
    return {
        "as_of_date": as_of_date,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "total_entries": len(entries),
        "suspense_account": has_suspense,
        "suspense_amount": difference if has_suspense else 0,
        "cost_center": cost_center
    }

@api_router.get("/reports/profit-loss")
async def get_profit_loss(start_date: str, end_date: str, cost_center: Optional[str] = None):
    """Generate Profit & Loss Statement"""
    query = {"posting_date": {"$gte": start_date, "$lte": end_date}}
    if cost_center:
        query["cost_center"] = cost_center
    
    entries = await db.journal_entries.find(query, {"_id": 0}).to_list(10000)
    
    revenue = {}
    expenses = {}
    
    for entry in entries:
        account = entry["account"]
        if "Revenue" in account or "Income" in account or "Sales" in account:
            if account not in revenue:
                revenue[account] = 0
            revenue[account] += entry["credit"] - entry["debit"]
        elif "Expense" in account or "Cost" in account or "Purchases" in account:
            if account not in expenses:
                expenses[account] = 0
            expenses[account] += entry["debit"] - entry["credit"]
    
    total_revenue = sum(revenue.values())
    total_expenses = sum(expenses.values())
    
    return {
        "period": {"start": start_date, "end": end_date},
        "revenue": revenue,
        "total_revenue": total_revenue,
        "expenses": expenses,
        "total_expenses": total_expenses,
        "net_profit": total_revenue - total_expenses,
        "cost_center": cost_center
    }

@api_router.get("/reports/trial-balance")
async def get_trial_balance(as_of_date: Optional[str] = None):
    """Generate Trial Balance"""
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).date().isoformat()
    
    entries = await db.journal_entries.find(
        {"posting_date": {"$lte": as_of_date}},
        {"_id": 0}
    ).to_list(10000)
    
    balances = {}
    for entry in entries:
        account = entry["account"]
        if account not in balances:
            balances[account] = {"debit": 0, "credit": 0}
        balances[account]["debit"] += entry["debit"]
        balances[account]["credit"] += entry["credit"]
    
    total_debit = sum(b["debit"] for b in balances.values())
    total_credit = sum(b["credit"] for b in balances.values())
    
    return {
        "as_of_date": as_of_date,
        "accounts": balances,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "difference": total_debit - total_credit,
        "in_balance": abs(total_debit - total_credit) < 0.01
    }

# PDF Export
@api_router.get("/reports/balance-sheet/pdf")
async def download_balance_sheet_pdf(as_of_date: Optional[str] = None):
    """Download Balance Sheet as PDF"""
    data = await get_balance_sheet(as_of_date)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"<b>Balance Sheet - Kairos Accounting</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Date
    date_para = Paragraph(f"As of: {data['as_of_date']}", styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 12))
    
    # Assets Table
    if data['assets']:
        assets_data = [["Assets", "Amount"]]
        for acc, bal in data['assets'].items():
            assets_data.append([acc, f"₹{bal['debit'] - bal['credit']:.2f}"])
        
        t = Table(assets_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(content=buffer.read(), media_type="application/pdf", 
                   headers={"Content-Disposition": f"attachment; filename=balance_sheet_{as_of_date}.pdf"})

# Excel Export
@api_router.get("/reports/trial-balance/excel")
async def download_trial_balance_excel(as_of_date: Optional[str] = None):
    """Download Trial Balance as Excel"""
    data = await get_trial_balance(as_of_date)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    # Header
    ws['A1'] = "Kairos Accounting - Trial Balance"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"As of: {data['as_of_date']}"
    
    # Column headers
    ws['A4'] = "Account"
    ws['B4'] = "Debit"
    ws['C4'] = "Credit"
    
    for cell in ['A4', 'B4', 'C4']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    row = 5
    for acc, bal in data['accounts'].items():
        ws[f'A{row}'] = acc
        ws[f'B{row}'] = bal['debit']
        ws[f'C{row}'] = bal['credit']
        row += 1
    
    # Totals
    ws[f'A{row}'] = "Total"
    ws[f'B{row}'] = data['total_debit']
    ws[f'C{row}'] = data['total_credit']
    ws[f'A{row}'].font = Font(bold=True)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(content=buffer.read(), 
                   media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": f"attachment; filename=trial_balance_{as_of_date}.xlsx"})

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Kairos Accounting - Storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()