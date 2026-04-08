# Kairos Accounting - Schedule III Financial Statements
# Companies Act 2013 - Division I compliant BS & P&L

from fastapi import APIRouter, HTTPException, Response
from datetime import datetime, timezone
from typing import Optional
from io import BytesIO
import uuid
import json

router = APIRouter(prefix="/financial-statements", tags=["financial-statements"])
db = None

def set_db(database):
    global db
    db = database

# ═══════════════════════════════════════════════════════
# DYNAMIC SCHEDULE III CLASSIFICATION
# Uses CoA category/sub_category instead of hardcoded names
# ═══════════════════════════════════════════════════════

COMPANY_NAME = "PolyMerx Specialty Chemicals Pvt. Ltd."

async def _get_company_name():
    """Fetch company name from DB settings, fallback to default."""
    if db is not None:
        try:
            settings = await db.company_settings.find_one({}, {"_id": 0, "legal_name": 1, "short_name": 1})
            if settings:
                return settings.get("legal_name") or settings.get("short_name") or COMPANY_NAME
        except Exception:
            pass
    return COMPANY_NAME

# Keywords to sub-classify within categories
BS_CLASSIFY = {
    "share_capital": lambda n: "share capital" in n.lower(),
    "reserves": lambda n: "retained" in n.lower() or "reserves" in n.lower() or "surplus" in n.lower(),
    "ppe": lambda n: "plant" in n.lower() or "machinery" in n.lower() or "equipment" in n.lower(),
    "acc_dep": lambda n: "accumulated depreciation" in n.lower(),
    "cwip": lambda n: "capital work" in n.lower(),
    "rou": lambda n: "right-of-use" in n.lower() or "rou" in n.lower(),
    "dta": lambda n: "deferred tax asset" in n.lower(),
    "dtl": lambda n: "deferred tax liab" in n.lower(),
    "trade_payables": lambda n: "accounts payable" in n.lower() or "trade payable" in n.lower(),
    "trade_receivables": lambda n: "accounts receivable" in n.lower() or "trade receivable" in n.lower(),
    "cash": lambda n: "cash" in n.lower() or "bank" in n.lower() or "fixed deposit" in n.lower(),
    "inventory": lambda n: "inventory" in n.lower() or "work-in-progress" in n.lower(),
    "gst_input": lambda n: "gst input" in n.lower(),
    "gst_output": lambda n: "gst output" in n.lower() or "gst (advance)" in n.lower(),
    "salary_payable": lambda n: "salary payable" in n.lower(),
    "provision": lambda n: "provision" in n.lower() or "warranty" in n.lower(),
    "lease_liability": lambda n: "lease liab" in n.lower(),
    "borrowings": lambda n: "loan" in n.lower() or "borrowing" in n.lower(),
}

PL_CLASSIFY = {
    "revenue_ops": lambda n, c: c == "Revenue" and ("sales revenue" in n.lower() or "export revenue" in n.lower()),
    "other_income": lambda n, c: c == "Revenue" and not ("sales revenue" in n.lower() or "export revenue" in n.lower()),
    "materials": lambda n, c: c == "Expense" and ("raw material" in n.lower() or "consumable" in n.lower()),
    "cogs": lambda n, c: "cost of goods" in n.lower(),
    "employee": lambda n, c: "employee" in (c or "").lower() or "salary" in n.lower() or "pf expense" in n.lower() or "esi expense" in n.lower() or "bonus" in n.lower() or "gratuity" in n.lower(),
    "finance": lambda n, c: "finance" in (c or "").lower() or "interest expense" in n.lower() or "lease interest" in n.lower() or "forex" in n.lower() or "bank charge" in n.lower(),
    "depreciation": lambda n, c: "depreciation" in n.lower() or "amortization" in n.lower() or "amortisation" in n.lower() or "rou asset" in n.lower(),
    "tax": lambda n, c: "tax" in (c or "").lower() and "expense" in n.lower(),
}


async def get_accounts_with_categories() -> list:
    """Get all CoA accounts with category info"""
    return await db.chart_of_accounts.find({}, {"_id": 0}).to_list(1000)

async def get_account_balances_map() -> dict:
    """Get all CoA balances as a map"""
    accounts = await db.chart_of_accounts.find({}, {"_id": 0}).to_list(1000)
    return {a["ledger_name"]: a.get("current_balance", 0) for a in accounts}


# ═══════════════════════════════════════════════════════
# BALANCE SHEET - Schedule III Format
# ═══════════════════════════════════════════════════════
@router.get("/balance-sheet")
async def get_balance_sheet(as_of_date: Optional[str] = None):
    """Generate Balance Sheet per Schedule III Companies Act 2013"""
    accounts = await get_accounts_with_categories()
    balances = {a["ledger_name"]: a.get("current_balance", 0) for a in accounts}

    # Classify accounts dynamically
    share_capital_items, reserves_items = [], []
    lt_borrowings_items, lt_provisions_items, other_lt_liab_items = [], [], []
    trade_pay_items, other_cl_items, st_provisions_items = [], [], []
    ppe_items, acc_dep_items, cwip_items, rou_items = [], [], [], []
    dta_items, other_nca_items = [], []
    inventory_items, receivable_items, cash_items = [], [], []
    gst_input_items, other_ca_items = [], []

    for a in accounts:
        name = a["ledger_name"]
        cat = a.get("category", "")
        sub = a.get("sub_category", "")
        bal = a.get("current_balance", 0)

        if cat == "Equity":
            if BS_CLASSIFY["share_capital"](name):
                share_capital_items.append((name, bal))
            else:
                reserves_items.append((name, bal))
        elif cat == "Liability":
            if sub == "Non-Current Liability":
                if BS_CLASSIFY["borrowings"](name):
                    lt_borrowings_items.append((name, bal))
                elif BS_CLASSIFY["lease_liability"](name):
                    other_lt_liab_items.append((name, bal))
                elif BS_CLASSIFY["provision"](name):
                    lt_provisions_items.append((name, bal))
                else:
                    other_lt_liab_items.append((name, bal))
            else:  # Current Liability
                if BS_CLASSIFY["trade_payables"](name):
                    trade_pay_items.append((name, bal))
                elif BS_CLASSIFY["provision"](name):
                    st_provisions_items.append((name, bal))
                else:
                    other_cl_items.append((name, bal))
        elif cat == "Asset":
            if sub == "Non-Current Asset":
                if BS_CLASSIFY["ppe"](name):
                    ppe_items.append((name, bal))
                elif BS_CLASSIFY["acc_dep"](name):
                    acc_dep_items.append((name, bal))
                elif BS_CLASSIFY["cwip"](name):
                    cwip_items.append((name, bal))
                elif BS_CLASSIFY["rou"](name):
                    rou_items.append((name, bal))
                elif BS_CLASSIFY["dta"](name):
                    dta_items.append((name, bal))
                else:
                    other_nca_items.append((name, bal))
            else:  # Current Asset
                if BS_CLASSIFY["inventory"](name):
                    inventory_items.append((name, bal))
                elif BS_CLASSIFY["trade_receivables"](name):
                    receivable_items.append((name, bal))
                elif BS_CLASSIFY["cash"](name):
                    cash_items.append((name, bal))
                elif BS_CLASSIFY["gst_input"](name):
                    gst_input_items.append((name, bal))
                else:
                    other_ca_items.append((name, bal))

    # Calculate P&L for current period to add to reserves
    revenue_total = sum(-a.get("current_balance", 0) for a in accounts if a.get("category") == "Revenue")
    expense_total = sum(a.get("current_balance", 0) for a in accounts if a.get("category") == "Expense")
    current_year_pl = revenue_total - expense_total

    # Credit-normal sum: negate balance (liab/equity have negative bal = positive amount)
    def sum_credit(items):
        return sum(-b for _, b in items)

    share_capital = sum_credit(share_capital_items)
    reserves_surplus = sum_credit(reserves_items) + current_year_pl
    shareholders_funds = share_capital + reserves_surplus

    long_term_borrowings = sum_credit(lt_borrowings_items)
    lt_provisions = sum_credit(lt_provisions_items)
    other_lt_liab = sum_credit(other_lt_liab_items)
    non_current_liabilities = long_term_borrowings + lt_provisions + other_lt_liab

    trade_payables = sum_credit(trade_pay_items)
    other_cl = sum_credit(other_cl_items)
    st_provisions = sum_credit(st_provisions_items)
    current_liabilities = trade_payables + other_cl + st_provisions

    total_equity_liabilities = shareholders_funds + non_current_liabilities + current_liabilities

    # Assets
    ppe_gross = sum(b for _, b in ppe_items)
    acc_dep = sum(abs(b) for _, b in acc_dep_items)
    ppe_net = ppe_gross - acc_dep
    capital_wip = sum(b for _, b in cwip_items)
    rou_asset = sum(b for _, b in rou_items)
    dta = sum(b for _, b in dta_items)
    other_nca = sum(b for _, b in other_nca_items)
    non_current_assets = ppe_net + capital_wip + rou_asset + dta + other_nca

    inventories = sum(b for _, b in inventory_items)
    trade_receivables = sum(b for _, b in receivable_items)
    cash_equivalents = sum(b for _, b in cash_items)
    gst_input = sum(b for _, b in gst_input_items)
    other_ca = sum(b for _, b in other_ca_items)
    current_assets = inventories + trade_receivables + cash_equivalents + gst_input + other_ca

    total_assets = non_current_assets + current_assets

    return {
        "report_type": "Balance Sheet",
        "format": "Schedule III - Companies Act 2013 (Division I)",
        "company_name": await _get_company_name(),
        "as_of_date": as_of_date or datetime.now(timezone.utc).date().isoformat(),
        "currency": "INR",

        "equity_and_liabilities": {
            "shareholders_funds": {
                "label": "1. Shareholders' Funds",
                "share_capital": {"label": "(a) Share Capital", "amount": round(share_capital, 2), "note": 1},
                "reserves_and_surplus": {"label": "(b) Reserves and Surplus", "amount": round(reserves_surplus, 2), "note": 2,
                                         "details": {"retained_earnings_opening": round(reserves_surplus - current_year_pl, 2),
                                                    "current_year_profit_loss": round(current_year_pl, 2)}},
                "total": round(shareholders_funds, 2)
            },
            "non_current_liabilities": {
                "label": "3. Non-current Liabilities",
                "long_term_borrowings": {"label": "(a) Long-term Borrowings", "amount": round(long_term_borrowings, 2), "note": 3},
                "other_long_term_liabilities": {"label": "(b) Other Long-term Liabilities", "amount": round(other_lt_liab, 2)},
                "long_term_provisions": {"label": "(c) Long-term Provisions", "amount": round(lt_provisions, 2)},
                "total": round(non_current_liabilities, 2)
            },
            "current_liabilities": {
                "label": "4. Current Liabilities",
                "trade_payables": {"label": "(b) Trade Payables", "amount": round(trade_payables, 2), "note": 4},
                "other_current_liabilities": {"label": "(c) Other Current Liabilities", "amount": round(other_cl, 2), "note": 5},
                "short_term_provisions": {"label": "(d) Short-term Provisions", "amount": round(st_provisions, 2)},
                "total": round(current_liabilities, 2)
            },
            "total": round(total_equity_liabilities, 2)
        },

        "assets": {
            "non_current_assets": {
                "label": "1. Non-current Assets",
                "property_plant_equipment": {
                    "label": "(a) Property, Plant and Equipment",
                    "gross_block": round(ppe_gross, 2),
                    "accumulated_depreciation": round(acc_dep, 2),
                    "net_block": round(ppe_net, 2),
                    "note": 6
                },
                "capital_wip": {"label": "Capital Work-in-Progress", "amount": round(capital_wip, 2)},
                "rou_asset": {"label": "Right-of-Use Asset", "amount": round(rou_asset, 2)},
                "deferred_tax_assets": {"label": "(c) Deferred Tax Assets (Net)", "amount": round(dta, 2)},
                "other_non_current_assets": {"label": "(d) Other Non-current Assets", "amount": round(other_nca, 2)},
                "total": round(non_current_assets, 2)
            },
            "current_assets": {
                "label": "2. Current Assets",
                "inventories": {"label": "(b) Inventories", "amount": round(inventories, 2), "note": 8,
                                "details": {n: round(b, 2) for n, b in inventory_items}},
                "trade_receivables": {"label": "(c) Trade Receivables", "amount": round(trade_receivables, 2), "note": 9},
                "cash_and_equivalents": {"label": "(d) Cash and Cash Equivalents", "amount": round(cash_equivalents, 2), "note": 10,
                                         "details": {n: round(b, 2) for n, b in cash_items}},
                "short_term_loans_advances": {"label": "(e) Short-term Loans and Advances", "amount": round(gst_input, 2)},
                "other_current_assets": {"label": "(f) Other Current Assets", "amount": round(other_ca, 2),
                                         "details": {n: round(b, 2) for n, b in other_ca_items}},
                "total": round(current_assets, 2)
            },
            "total": round(total_assets, 2)
        },

        "is_balanced": abs(total_equity_liabilities - total_assets) < 1,
        "difference": round(total_equity_liabilities - total_assets, 2)
    }


# ═══════════════════════════════════════════════════════
# STATEMENT OF PROFIT & LOSS - Schedule III Format
# ═══════════════════════════════════════════════════════
@router.get("/profit-and-loss")
async def get_profit_and_loss(start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Generate Statement of Profit & Loss per Schedule III Companies Act 2013"""
    accounts = await get_accounts_with_categories()

    # Classify accounts dynamically
    revenue_ops, other_income_items = [], []
    materials_items, cogs_items, employee_items = [], [], []
    finance_items, depreciation_items, tax_items = [], [], []
    other_expense_items = []

    for a in accounts:
        name = a["ledger_name"]
        cat = a.get("category", "")
        sub = a.get("sub_category", "")
        bal = a.get("current_balance", 0)
        if bal == 0:
            continue

        if PL_CLASSIFY["revenue_ops"](name, cat):
            revenue_ops.append((name, abs(bal)))
        elif PL_CLASSIFY["other_income"](name, cat):
            other_income_items.append((name, abs(bal)))
        elif PL_CLASSIFY["cogs"](name, cat):
            cogs_items.append((name, bal))
        elif PL_CLASSIFY["materials"](name, cat):
            materials_items.append((name, bal))
        elif PL_CLASSIFY["employee"](name, sub):
            employee_items.append((name, bal))
        elif PL_CLASSIFY["tax"](name, sub):
            tax_items.append((name, bal))
        elif PL_CLASSIFY["depreciation"](name, cat):
            depreciation_items.append((name, bal))
        elif PL_CLASSIFY["finance"](name, sub):
            finance_items.append((name, bal))
        elif cat == "Expense":
            other_expense_items.append((name, bal))

    revenue_operations = sum(b for _, b in revenue_ops)
    other_income = sum(b for _, b in other_income_items)
    total_revenue = revenue_operations + other_income

    cost_materials = sum(b for _, b in materials_items)
    cogs = sum(b for _, b in cogs_items)
    employee_benefits = sum(b for _, b in employee_items)
    finance_costs = sum(b for _, b in finance_items)
    depreciation = sum(b for _, b in depreciation_items)
    other_expenses = sum(b for _, b in other_expense_items)
    current_tax = sum(b for _, b in tax_items)

    total_expenses = cost_materials + cogs + employee_benefits + finance_costs + depreciation + other_expenses

    profit_before_tax = total_revenue - total_expenses
    profit_for_period = profit_before_tax - current_tax

    return {
        "report_type": "Statement of Profit and Loss",
        "format": "Schedule III - Companies Act 2013 (Division I)",
        "company_name": await _get_company_name(),
        "period": {
            "from": start_date or "2025-04-01",
            "to": end_date or datetime.now(timezone.utc).date().isoformat()
        },
        "currency": "INR",

        "line_items": [
            {"sl": "I", "particular": "Revenue from Operations", "amount": round(revenue_operations, 2), "note": 11,
             "details": [{"account": n, "amount": round(b, 2)} for n, b in revenue_ops]},

            {"sl": "II", "particular": "Other Income", "amount": round(other_income, 2), "note": 12,
             "details": [{"account": n, "amount": round(b, 2)} for n, b in other_income_items]},

            {"sl": "III", "particular": "Total Revenue (I + II)", "amount": round(total_revenue, 2), "is_total": True},

            {"sl": "IV", "particular": "Expenses:", "is_header": True},
            {"sl": "", "particular": "Cost of Materials Consumed", "amount": round(cost_materials, 2), "note": 13},
            {"sl": "", "particular": "Cost of Goods Sold", "amount": round(cogs, 2),
             "details": [{"account": n, "amount": round(b, 2)} for n, b in cogs_items]},
            {"sl": "", "particular": "Employee Benefits Expense", "amount": round(employee_benefits, 2), "note": 14,
             "details": [{"account": n, "amount": round(b, 2)} for n, b in employee_items]},
            {"sl": "", "particular": "Finance Costs", "amount": round(finance_costs, 2), "note": 15,
             "details": [{"account": n, "amount": round(b, 2)} for n, b in finance_items]},
            {"sl": "", "particular": "Depreciation and Amortisation Expense", "amount": round(depreciation, 2), "note": 6,
             "details": [{"account": n, "amount": round(b, 2)} for n, b in depreciation_items]},
            {"sl": "", "particular": "Other Expenses", "amount": round(other_expenses, 2), "note": 16,
             "details": [{"account": n, "amount": round(b, 2)} for n, b in other_expense_items]},
            {"sl": "", "particular": "Total Expenses (IV)", "amount": round(total_expenses, 2), "is_total": True},

            {"sl": "V", "particular": "Profit before Exceptional and Extraordinary Items and Tax (III - IV)", "amount": round(profit_before_tax, 2)},
            {"sl": "VI", "particular": "Exceptional Items", "amount": 0},
            {"sl": "IX", "particular": "Profit before Tax", "amount": round(profit_before_tax, 2), "is_total": True},

            {"sl": "X", "particular": "Tax Expense:", "is_header": True},
            {"sl": "", "particular": "(1) Current Tax / Income Tax", "amount": round(current_tax, 2),
             "details": [{"account": n, "amount": round(b, 2)} for n, b in tax_items]},

            {"sl": "XI", "particular": "Profit (Loss) for the Period", "amount": round(profit_for_period, 2), "is_total": True, "is_final": True},
        ],

        "summary": {
            "total_revenue": round(total_revenue, 2),
            "total_expenses": round(total_expenses, 2),
            "profit_before_tax": round(profit_before_tax, 2),
            "tax_expense": round(current_tax, 2),
            "net_profit": round(profit_for_period, 2),
            "gross_margin_pct": round((total_revenue - cost_materials - cogs) / total_revenue * 100, 2) if total_revenue > 0 else 0,
        }
    }


# ═══════════════════════════════════════════════════════
# TRIAL BALANCE (Enhanced)
# ═══════════════════════════════════════════════════════
@router.get("/trial-balance")
async def get_trial_balance(as_of_date: Optional[str] = None):
    accounts = await db.chart_of_accounts.find({}, {"_id": 0}).sort("ledger_name", 1).to_list(1000)

    entries = []
    total_debit = 0
    total_credit = 0

    for acc in accounts:
        bal = acc.get("current_balance", 0)
        if bal == 0:
            continue
        debit = bal if bal > 0 else 0
        credit = abs(bal) if bal < 0 else 0
        entries.append({
            "account": acc["ledger_name"],
            "category": acc.get("category", ""),
            "debit": round(debit, 2),
            "credit": round(credit, 2),
        })
        total_debit += debit
        total_credit += credit

    return {
        "report_type": "Trial Balance",
        "company_name": await _get_company_name(),
        "as_of_date": as_of_date or datetime.now(timezone.utc).date().isoformat(),
        "entries": entries,
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "difference": round(total_debit - total_credit, 2),
        "in_balance": abs(total_debit - total_credit) < 1
    }



# ═══════════════════════════════════════════════════════
# PDF & EXCEL EXPORTS
# ═══════════════════════════════════════════════════════
@router.get("/balance-sheet/export/excel")
async def export_bs_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    data = await get_balance_sheet()
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 18

    header_font = Font(name='Arial', bold=True, size=12)
    sub_font = Font(name='Arial', size=10, italic=True)
    bold_font = Font(name='Arial', bold=True, size=10)
    num_font = Font(name='Arial', size=10)
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(bottom=Side(style='thin'))

    ws.append([data["company_name"]])
    ws.merge_cells('A1:C1')
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([f"Balance Sheet as at {data['as_of_date']}"])
    ws.merge_cells('A2:C2')
    ws['A2'].font = sub_font
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([data["format"]])
    ws.merge_cells('A3:C3')
    ws['A3'].font = Font(size=8, italic=True)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(["Particulars", "Note", "Amount (INR)"])
    for cell in ws[5]:
        cell.font = bold_font
        cell.border = Border(bottom=Side(style='double'))

    def add_section(title, items, total_label, total_amount):
        ws.append([title])
        ws[ws.max_row][0].font = bold_font
        for label, amount, note in items:
            row = ws.max_row + 1
            ws.append([f"  {label}", note or "", amount])
            ws[row][2].number_format = '#,##0.00'
            ws[row][2].font = num_font
        row = ws.max_row + 1
        ws.append([f"  {total_label}", "", total_amount])
        ws[row][0].font = bold_font
        ws[row][2].font = bold_font
        ws[row][2].number_format = '#,##0.00'
        for cell in ws[row]:
            cell.fill = total_fill
            cell.border = thin_border

    el = data["equity_and_liabilities"]
    ws.append(["I. EQUITY AND LIABILITIES"])
    ws[ws.max_row][0].font = Font(name='Arial', bold=True, size=11)

    add_section("1. Shareholders' Funds", [
        ("(a) Share Capital", el["shareholders_funds"]["share_capital"]["amount"], 1),
        ("(b) Reserves and Surplus", el["shareholders_funds"]["reserves_and_surplus"]["amount"], 2),
    ], "Total Shareholders' Funds", el["shareholders_funds"]["total"])

    add_section("3. Non-current Liabilities", [
        ("(a) Long-term Borrowings", el["non_current_liabilities"]["long_term_borrowings"]["amount"], 3),
    ], "Total Non-current Liabilities", el["non_current_liabilities"]["total"])

    add_section("4. Current Liabilities", [
        ("(b) Trade Payables", el["current_liabilities"]["trade_payables"]["amount"], 4),
        ("(c) Other Current Liabilities", el["current_liabilities"]["other_current_liabilities"]["amount"], 5),
    ], "Total Current Liabilities", el["current_liabilities"]["total"])

    row = ws.max_row + 1
    ws.append(["TOTAL EQUITY & LIABILITIES", "", el["total"]])
    ws[row][0].font = Font(name='Arial', bold=True, size=11)
    ws[row][2].font = Font(name='Arial', bold=True, size=11)
    ws[row][2].number_format = '#,##0.00'
    for cell in ws[row]:
        cell.border = Border(top=Side(style='double'), bottom=Side(style='double'))

    ws.append([])
    a = data["assets"]
    ws.append(["II. ASSETS"])
    ws[ws.max_row][0].font = Font(name='Arial', bold=True, size=11)

    nca = a["non_current_assets"]
    nca_items = [
        ("(a) Property, Plant & Equipment (Net)", nca["property_plant_equipment"]["net_block"], 6),
    ]
    if nca.get("capital_wip", {}).get("amount", 0) > 0:
        nca_items.append(("Capital Work-in-Progress", nca["capital_wip"]["amount"], ""))
    if nca.get("rou_asset", {}).get("amount", 0) > 0:
        nca_items.append(("Right-of-Use Asset", nca["rou_asset"]["amount"], ""))
    if nca.get("deferred_tax_assets", {}).get("amount", 0) > 0:
        nca_items.append(("Deferred Tax Assets (Net)", nca["deferred_tax_assets"]["amount"], ""))
    if nca.get("other_non_current_assets", {}).get("amount", 0) > 0:
        nca_items.append(("Other Non-current Assets", nca["other_non_current_assets"]["amount"], 7))
    add_section("1. Non-current Assets", nca_items, "Total Non-current Assets", nca["total"])

    add_section("2. Current Assets", [
        ("(b) Inventories", a["current_assets"]["inventories"]["amount"], 8),
        ("(c) Trade Receivables", a["current_assets"]["trade_receivables"]["amount"], 9),
        ("(d) Cash and Cash Equivalents", a["current_assets"]["cash_and_equivalents"]["amount"], 10),
        ("(e) Short-term Loans & Advances", a["current_assets"]["short_term_loans_advances"]["amount"], ""),
    ], "Total Current Assets", a["current_assets"]["total"])

    row = ws.max_row + 1
    ws.append(["TOTAL ASSETS", "", a["total"]])
    ws[row][0].font = Font(name='Arial', bold=True, size=11)
    ws[row][2].font = Font(name='Arial', bold=True, size=11)
    ws[row][2].number_format = '#,##0.00'
    for cell in ws[row]:
        cell.border = Border(top=Side(style='double'), bottom=Side(style='double'))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": "attachment; filename=Balance_Sheet_Schedule_III.xlsx"})


@router.get("/profit-and-loss/export/excel")
async def export_pl_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    data = await get_profit_and_loss()
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18

    ws.append([data["company_name"]])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(name='Arial', bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([f"Statement of Profit and Loss for {data['period']['from']} to {data['period']['to']}"])
    ws.merge_cells('A2:D2')
    ws['A2'].font = Font(name='Arial', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(["Sl.", "Particulars", "Note", "Amount (INR)"])
    for cell in ws[4]:
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.border = Border(bottom=Side(style='double'))

    for item in data["line_items"]:
        row_num = ws.max_row + 1
        ws.append([item.get("sl", ""), item.get("particular", ""), item.get("note", ""),
                   item.get("amount") if not item.get("is_header") else ""])
        if item.get("is_total") or item.get("is_final"):
            for cell in ws[row_num]:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.border = Border(top=Side(style='thin'))
        if item.get("is_final"):
            for cell in ws[row_num]:
                cell.border = Border(top=Side(style='double'), bottom=Side(style='double'))
        if ws[row_num][3].value is not None:
            ws[row_num][3].number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": "attachment; filename=Profit_Loss_Schedule_III.xlsx"})


@router.get("/trial-balance/export/excel")
async def export_tb_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    data = await get_trial_balance()
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws.append([data["company_name"]])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(name='Arial', bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([f"Trial Balance as at {data['as_of_date']}"])
    ws.merge_cells('A2:D2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(["Account", "Category", "Debit (INR)", "Credit (INR)"])
    for cell in ws[4]:
        cell.font = Font(name='Arial', bold=True)
        cell.border = Border(bottom=Side(style='double'))

    for entry in data["entries"]:
        row = ws.max_row + 1
        ws.append([entry["account"], entry["category"], entry["debit"], entry["credit"]])
        ws[row][2].number_format = '#,##0.00'
        ws[row][3].number_format = '#,##0.00'

    row = ws.max_row + 1
    ws.append(["TOTAL", "", data["total_debit"], data["total_credit"]])
    for cell in ws[row]:
        cell.font = Font(name='Arial', bold=True)
        cell.border = Border(top=Side(style='double'), bottom=Side(style='double'))
    ws[row][2].number_format = '#,##0.00'
    ws[row][3].number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": "attachment; filename=Trial_Balance.xlsx"})
