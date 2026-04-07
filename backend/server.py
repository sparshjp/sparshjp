from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Header, Query
from fastapi.responses import Response, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import requests
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
import base64
import io
from PIL import Image
import csv
import json
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import routes_auth

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Object Storage Configuration
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "kairos-accounting"
storage_key = None

# Initialize storage
def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str) -> tuple:
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# Create the main app
app = FastAPI(title="Kairos Accounting")

# Create API router
api_router = APIRouter(prefix="/api")

# Pydantic Models
class ChartOfAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ledger_name: str
    category: str
    opening_balance: float = 0.0
    current_balance: float = 0.0
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CostCenter(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TransactionBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_prompt: Optional[str] = None
    user_id: str = "default_user"
    module: str
    status: str = "draft"
    posting_date: Optional[str] = None
    service_period_start: Optional[str] = None
    service_period_end: Optional[str] = None
    business_unit: Optional[str] = None
    cost_center: str = "General"
    extracted_data: Optional[Dict[str, Any]] = None
    journal_entries: Optional[List[Dict[str, Any]]] = None
    document_id: Optional[str] = None
    import_source: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    posted_at: Optional[str] = None

class DocumentUpload(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    storage_path: str
    original_filename: str
    content_type: str
    size: int
    extracted_data: Optional[Dict[str, Any]] = None
    transaction_id: Optional[str] = None
    is_deleted: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class JournalEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_id: str
    account: str
    debit: float = 0.0
    credit: float = 0.0
    description: str
    posting_date: str
    cost_center: str = "General"
    tax_details: Optional[Dict[str, Any]] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class VendorClient(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str
    name: str
    gstin: Optional[str] = None
    pan: Optional[str] = None
    legal_name: Optional[str] = None
    constitution: Optional[str] = None
    status: str = "Active"
    address: Optional[str] = None
    state: Optional[str] = None
    state_code: Optional[str] = None
    contact: Optional[str] = None
    email: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_name: str
    hsn_sac: Optional[str] = None
    hsn_code: Optional[str] = None
    gst_rate: float = 18.0
    unit: str = "pcs"
    current_stock: float = 0.0
    landed_cost: float = 0.0
    valuation_method: str = "FIFO"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PromptRequest(BaseModel):
    prompt: str
    module: str
    user_id: str = "default_user"
    document_id: Optional[str] = None
    cost_center: str = "General"

class PostTransactionRequest(BaseModel):
    transaction_id: str
    cost_center: Optional[str] = None

class ReportQuery(BaseModel):
    query: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class CSVImportRequest(BaseModel):
    module: str
    csv_data: str

# GSTIN/PAN Validation & Intelligence
GSTIN_STATE_CODES = {
    "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan",
    "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh",
    "13": "Nagaland", "14": "Manipur", "15": "Mizoram", "16": "Tripura",
    "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
    "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
    "25": "Daman & Diu", "26": "Dadra & Nagar Haveli", "27": "Maharashtra",
    "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
    "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman & Nicobar", "36": "Telangana",
    "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory",
}

PAN_ENTITY_MAP = {
    "A": "Association of Persons (AOP)", "B": "Body of Individuals (BOI)",
    "C": "Company", "F": "Firm / LLP", "G": "Government",
    "H": "Hindu Undivided Family (HUF)", "J": "Artificial Juridical Person",
    "L": "Local Authority", "P": "Individual / Proprietor", "T": "Trust",
}

import re

def validate_gstin(gstin: str) -> Dict[str, Any]:
    """Validate GSTIN format, extract PAN, state, entity type"""
    gstin = gstin.strip().upper()
    result = {"valid": False, "gstin": gstin, "errors": []}

    if len(gstin) != 15:
        result["errors"].append(f"GSTIN must be 15 characters, got {len(gstin)}")
        return result

    pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[0-9A-Z]{1}[Z]{1}[0-9A-Z]{1}$'
    if not re.match(pattern, gstin):
        result["errors"].append("Invalid GSTIN format")
        return result

    state_code = gstin[:2]
    pan = gstin[2:12]
    entity_char = pan[3]

    state_name = GSTIN_STATE_CODES.get(state_code)
    if not state_name:
        result["errors"].append(f"Invalid state code: {state_code}")
        return result

    result["valid"] = True
    result["pan"] = pan
    result["state_code"] = state_code
    result["state_name"] = state_name
    result["entity_type"] = PAN_ENTITY_MAP.get(entity_char, "Unknown")
    result["entity_number"] = gstin[12]
    return result

def validate_pan(pan: str) -> Dict[str, Any]:
    """Validate PAN format and extract entity type"""
    pan = pan.strip().upper()
    result = {"valid": False, "pan": pan, "errors": []}

    if len(pan) != 10:
        result["errors"].append(f"PAN must be 10 characters, got {len(pan)}")
        return result

    pattern = r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$'
    if not re.match(pattern, pan):
        result["errors"].append("Invalid PAN format (AAAAA9999A)")
        return result

    entity_char = pan[3]
    result["valid"] = True
    result["entity_type"] = PAN_ENTITY_MAP.get(entity_char, "Unknown")
    return result

async def lookup_gstin(gstin: str) -> Dict[str, Any]:
    """Validate GSTIN and return extracted intelligence"""
    info = validate_gstin(gstin)
    if not info["valid"]:
        return {"legal_name": "", "constitution": "", "status": "Invalid GSTIN", "errors": info["errors"]}

    return {
        "legal_name": "",
        "constitution": info.get("entity_type", ""),
        "status": "Active",
        "state_name": info.get("state_name", ""),
        "state_code": info.get("state_code", ""),
        "pan": info.get("pan", ""),
        "gstin_valid": True
    }

# AI Services
async def parse_prompt_with_ai(prompt: str, module: str, extracted_ocr: Optional[Dict] = None, cost_center: str = "General") -> Dict[str, Any]:
    """Use Claude Sonnet 4.5 to parse user prompt and generate journal entries"""
    try:
        chat = LlmChat(
            api_key=EMERGENT_KEY,
            session_id=f"prompt-{uuid.uuid4()}",
            system_message="""You are an expert Indian ERP accountant for Kairos Accounting system. Parse the user's natural language prompt and extract:
            1. posting_date (format: YYYY-MM-DD)
            2. service_period_start (format: YYYY-MM-DD)
            3. service_period_end (format: YYYY-MM-DD)
            4. business_unit (e.g., Mumbai Office, Gujarat Plant)
            5. journal_entries: List of structured entries with:
               - account: Ledger name
               - debit: Amount (use 0 if credit entry)
               - credit: Amount (use 0 if debit entry)
               - description: Brief description
               - tax_details: {igst: 0, cgst: 0, sgst: 0, cess: 0, tds: 0}
            
            For GST transactions:
            - Determine if interstate (IGST) or intrastate (CGST + SGST)
            - Default GST rate: 18%
            - Include proper GST accounts: Input GST / Output GST
            - Calculate landed cost for goods = Purchase Price + IGST/CGST+SGST + Freight
            
            For TDS:
            - Identify salary payments and apply appropriate TDS
            - Include TDS Payable account
            
            Return ONLY a valid JSON object with these fields. No markdown, no explanations, no code fences."""
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")
        
        prompt_text = f"Module: {module}\nCost Center: {cost_center}\nPrompt: {prompt}"
        if extracted_ocr:
            prompt_text += f"\n\nExtracted OCR Data: {json.dumps(extracted_ocr)}"
        
        user_message = UserMessage(text=prompt_text)
        response = await chat.send_message(user_message)
        
        # Clean response: strip markdown code fences if present
        cleaned = response.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1] if "\n" in cleaned else cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()
        
        result = json.loads(cleaned)
        return result
    except Exception as e:
        logging.error(f"AI parsing error: {e}")
        return {
            "posting_date": datetime.now(timezone.utc).date().isoformat(),
            "service_period_start": datetime.now(timezone.utc).date().isoformat(),
            "service_period_end": datetime.now(timezone.utc).date().isoformat(),
            "business_unit": "Main Office",
            "journal_entries": []
        }

async def extract_document_data(image_data: bytes, filename: str) -> Dict[str, Any]:
    """Use Gemini 3 vision for OCR extraction"""
    try:
        img = Image.open(io.BytesIO(image_data))
        buffered = io.BytesIO()
        img.save(buffered, format="PNG")
        img_base64 = base64.b64encode(buffered.getvalue()).decode()
        
        chat = LlmChat(
            api_key=EMERGENT_KEY,
            session_id=f"ocr-{uuid.uuid4()}",
            system_message="""You are an OCR expert for Indian financial documents. Extract:
            - Vendor/Client name
            - GSTIN (15-character GST identification number)
            - Invoice number
            - Invoice date (format: YYYY-MM-DD)
            - Total amount
            - Line items with: item_name, hsn_code, quantity, rate, amount
            - GST breakdown: {cgst_rate, cgst_amount, sgst_rate, sgst_amount, igst_rate, igst_amount, cess}
            - TDS if applicable
            - Freight charges if any
            
            Return ONLY valid JSON. No markdown."""
        ).with_model("gemini", "gemini-3-flash-preview")
        
        user_message = UserMessage(
            text="Extract all data from this invoice/receipt image. Return JSON only.",
            file_contents=[ImageContent(image_base64=img_base64)]
        )
        response = await chat.send_message(user_message)
        
        # Clean response: strip markdown code fences if present
        cleaned = response.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1] if "\n" in cleaned else cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()
        
        result = json.loads(cleaned)
        return result
    except Exception as e:
        logging.error(f"OCR extraction error: {e}")
        return {
            "vendor_name": "Unknown",
            "gstin": "",
            "invoice_number": "",
            "invoice_date": "",
            "total_amount": 0,
            "line_items": [],
            "gst_breakdown": {}
        }

# CSV Validation and Processing
async def validate_csv_data(csv_data: str, module: str) -> Dict[str, Any]:
    """Validate CSV data against Zoho-standard headers and CoA"""
    required_headers = {
        "purchases": ["Date", "Entity Name", "Item/Service", "Rate", "GST Rate", "Total"],
        "sales": ["Date", "Entity Name", "Item/Service", "Rate", "GST Rate", "Total"],
        "journals": ["Date", "Ledger", "Debit", "Credit", "Description"],
        "payments": ["Date", "Entity Name", "Amount", "Payment Mode"]
    }
    
    errors = []
    warnings = []
    
    try:
        csv_file = StringIO(csv_data)
        reader = csv.DictReader(csv_file)
        headers = reader.fieldnames
        
        if not headers:
            return {"valid": False, "errors": ["Empty CSV or no headers found"], "warnings": [], "row_count": 0}
        
        # Check required headers
        module_key = module.replace("-", "")
        if module_key in required_headers:
            missing = [h for h in required_headers[module_key] if h not in headers]
            if missing:
                errors.extend([f"Missing required header: {h}" for h in missing])
        else:
            warnings.append(f"Unknown module '{module}'. No header validation applied.")
        
        # Load CoA ledgers for cross-validation
        coa_ledgers = set()
        coa_list = await db.chart_of_accounts.find({}, {"_id": 0, "ledger_name": 1}).to_list(1000)
        coa_ledgers = {a["ledger_name"] for a in coa_list}
        
        # Validate data rows
        rows = list(reader)
        if len(rows) == 0:
            warnings.append("CSV file has headers but no data rows")
        
        gstin_seen = set()
        for idx, row in enumerate(rows, start=2):
            # Validate date format
            if "Date" in row and row["Date"]:
                try:
                    datetime.strptime(row["Date"], "%Y-%m-%d")
                except ValueError:
                    errors.append(f"Row {idx}: Invalid date format '{row['Date']}'. Use YYYY-MM-DD")
            elif "Date" in row and not row["Date"]:
                errors.append(f"Row {idx}: Date is empty")
            
            # Validate ledger existence for journals
            if module == "journals" and "Ledger" in row and row["Ledger"]:
                if row["Ledger"] not in coa_ledgers:
                    warnings.append(f"Row {idx}: Ledger '{row['Ledger']}' not in Chart of Accounts (will not update CoA balance)")
            
            # Validate numeric fields
            for num_field in ["Rate", "Total", "Amount", "Debit", "Credit", "GST Rate"]:
                if num_field in row and row.get(num_field):
                    try:
                        float(row[num_field])
                    except ValueError:
                        errors.append(f"Row {idx}: '{num_field}' must be a number, got '{row[num_field]}'")
            
            # Validate GSTIN format in entity rows if present
            if "GSTIN" in row and row.get("GSTIN"):
                gstin_result = validate_gstin(row["GSTIN"])
                if not gstin_result["valid"]:
                    warnings.append(f"Row {idx}: Invalid GSTIN '{row['GSTIN']}' - {', '.join(gstin_result['errors'])}")
            
            # Validate entity exists for purchases/sales
            if module in ["purchases", "sales"] and "Entity Name" in row and row.get("Entity Name"):
                entity = await db.entities.find_one({"name": row["Entity Name"]}, {"_id": 0})
                if not entity:
                    warnings.append(f"Row {idx}: Entity '{row['Entity Name']}' not found in Master Data")
            
            # Check debit/credit balance for journals
            if module == "journals":
                debit = float(row.get("Debit", 0) or 0)
                credit = float(row.get("Credit", 0) or 0)
                if debit == 0 and credit == 0:
                    warnings.append(f"Row {idx}: Both Debit and Credit are zero")
        
        # For journals, check total debit == total credit
        if module == "journals" and len(rows) > 0:
            total_d = sum(float(r.get("Debit", 0) or 0) for r in rows)
            total_c = sum(float(r.get("Credit", 0) or 0) for r in rows)
            if abs(total_d - total_c) > 0.01:
                errors.append(f"Journal not balanced: Total Debit ({total_d:.2f}) != Total Credit ({total_c:.2f})")
        
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "row_count": len(rows),
            "coa_ledger_count": len(coa_ledgers)
        }
    except Exception as e:
        return {
            "valid": False,
            "errors": [f"CSV parsing error: {str(e)}"],
            "warnings": [],
            "row_count": 0
        }

# API Endpoints

@api_router.get("/health")
async def health_check():
    """System health check endpoint"""
    try:
        # Check database connection
        await db.command('ping')
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return {
        "status": "ok",
        "service": "kairos-accounting",
        "database": db_status,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

@api_router.get("/")
async def root():
    return {"message": "Kairos Accounting API", "version": "2.0.0"}

# Chart of Accounts
@api_router.post("/coa/upload")
async def upload_coa(file: UploadFile = File(...)):
    """Bulk upload Chart of Accounts from CSV"""
    try:
        content = await file.read()
        csv_data = content.decode('utf-8')
        csv_file = StringIO(csv_data)
        reader = csv.DictReader(csv_file)
        
        imported = 0
        errors = []
        
        for idx, row in enumerate(reader, start=2):
            try:
                coa = ChartOfAccount(
                    ledger_name=row.get("Ledger Name", ""),
                    category=row.get("Category", "Other"),
                    opening_balance=float(row.get("Opening Balance", 0))
                )
                coa_dict = coa.model_dump()
                coa_dict["current_balance"] = coa_dict["opening_balance"]
                await db.chart_of_accounts.insert_one(coa_dict)
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {"imported": imported, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/coa")
async def get_coa():
    """Get all Chart of Accounts"""
    coa_list = await db.chart_of_accounts.find({"is_active": {"$ne": False}}, {"_id": 0}).to_list(1000)
    return coa_list


@api_router.get("/ledgers")
async def get_ledgers_alias():
    """Alias for /coa endpoint - returns Chart of Accounts"""
    items = await db.chart_of_accounts.find({}, {"_id": 0}).to_list(500)
    return items

@api_router.post("/coa")
async def create_coa(coa: ChartOfAccount):
    """Create single Chart of Account entry"""
    coa_dict = coa.model_dump()
    coa_dict["current_balance"] = coa_dict["opening_balance"]
    await db.chart_of_accounts.insert_one(coa_dict)
    await audit_trail.log_audit(audit_trail.ACTION_CREATE, audit_trail.DOC_COA, coa_dict.get("ledger_name", ""), coa_dict.get("ledger_name", ""), snapshot=coa_dict, notes=f"Ledger created: {coa_dict.get('ledger_name')} ({coa_dict.get('category')})")
    return coa_dict

# Cost Centers
@api_router.post("/cost-centers")
async def create_cost_center(cc: CostCenter):
    """Create cost center"""
    cc_dict = cc.model_dump()
    await db.cost_centers.insert_one(cc_dict)
    await audit_trail.log_audit(audit_trail.ACTION_CREATE, audit_trail.DOC_COST_CENTER, cc_dict.get("name", ""), cc_dict.get("name", ""), snapshot=cc_dict, notes=f"Cost Center created: {cc_dict.get('name')}")
    return cc_dict

@api_router.get("/cost-centers")
async def get_cost_centers():
    """Get all cost centers"""
    centers = await db.cost_centers.find({"is_active": True}, {"_id": 0}).to_list(100)
    return centers

# Vendor/Client Master
@api_router.get("/validate/gstin/{gstin}")
async def api_validate_gstin(gstin: str):
    """Validate GSTIN format and extract PAN, state, entity type"""
    result = validate_gstin(gstin)
    return result

@api_router.get("/validate/pan/{pan}")
async def api_validate_pan(pan: str):
    """Validate PAN format and extract entity type"""
    result = validate_pan(pan)
    return result

@api_router.post("/entities")
async def create_entity(entity: VendorClient):
    """Create vendor or client with GSTIN/PAN intelligence"""
    entity_dict = entity.model_dump()
    
    # Auto-validate and enrich GSTIN if provided
    if entity.gstin:
        gstin_info = validate_gstin(entity.gstin)
        if gstin_info["valid"]:
            if not entity.pan:
                entity_dict["pan"] = gstin_info.get("pan", "")
            if not entity.constitution:
                entity_dict["constitution"] = gstin_info.get("entity_type", "")
            entity_dict["state_name"] = gstin_info.get("state_name", "")
            entity_dict["state_code"] = gstin_info.get("state_code", "")
            entity_dict["gstin_valid"] = True
            # Resolve canonical state info from GST rules engine
            import gst_rules
            gst_state_code = gst_rules.extract_state_from_gstin(entity.gstin)
            if gst_state_code:
                state_info = gst_rules.get_state_info(gst_state_code)
                if state_info:
                    entity_dict["gst_state_code"] = gst_state_code
                    entity_dict["state"] = state_info["name"]
        else:
            entity_dict["gstin_valid"] = False
            entity_dict["gstin_errors"] = gstin_info.get("errors", [])
    
    # Auto-validate PAN if provided
    if entity.pan:
        pan_info = validate_pan(entity.pan)
        if pan_info["valid"] and not entity.constitution:
            entity_dict["constitution"] = pan_info.get("entity_type", "")
    
    await db.entities.insert_one(entity_dict)
    del entity_dict["_id"]
    await audit_trail.log_audit(audit_trail.ACTION_CREATE, audit_trail.DOC_ENTITY, entity_dict.get("name", ""), entity_dict.get("name", ""), snapshot=entity_dict, notes=f"{entity_dict.get('entity_type', 'entity').title()} created: {entity_dict.get('name')}")
    return entity_dict

@api_router.get("/entities")
async def get_entities(entity_type: Optional[str] = None):
    """Get vendors or clients"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    entities = await db.entities.find(query, {"_id": 0}).to_list(1000)
    return entities

# CSV Import
@api_router.post("/import/validate")
async def validate_import(request: CSVImportRequest):
    """Validate CSV before import"""
    validation = await validate_csv_data(request.csv_data, request.module)
    return validation

@api_router.post("/import/process")
async def process_import(file: UploadFile = File(...), module: str = Query(...)):
    """Process CSV import for any module"""
    try:
        content = await file.read()
        csv_data = content.decode('utf-8')
        
        # Validate first
        validation = await validate_csv_data(csv_data, module)
        if not validation["valid"]:
            return {"success": False, "validation": validation}
        
        # Process import
        csv_file = StringIO(csv_data)
        reader = csv.DictReader(csv_file)
        imported = 0
        
        for row in reader:
            # Create transaction based on module
            txn = TransactionBase(
                module=module,
                posting_date=row.get("Date"),
                import_source="csv",
                status="draft"
            )
            
            # Build journal entries from CSV
            entries = []
            if module == "journals":
                entries.append({
                    "account": row.get("Ledger"),
                    "debit": float(row.get("Debit", 0)),
                    "credit": float(row.get("Credit", 0)),
                    "description": row.get("Description", "")
                })
            elif module in ["purchases", "sales"]:
                # Auto-generate journal entries from purchase/sale data
                total = float(row.get("Total", 0))
                gst_rate = float(row.get("GST Rate", 0))
                base = total / (1 + gst_rate / 100)
                gst_amt = total - base
                
                if module == "purchases":
                    entries.append({"account": "Purchases", "debit": base, "credit": 0, "description": row.get("Item/Service")})
                    entries.append({"account": "Input GST", "debit": gst_amt, "credit": 0, "description": "GST"})
                    entries.append({"account": row.get("Entity Name"), "debit": 0, "credit": total, "description": "Vendor"})
                else:
                    entries.append({"account": row.get("Entity Name"), "debit": total, "credit": 0, "description": "Customer"})
                    entries.append({"account": "Sales", "debit": 0, "credit": base, "description": row.get("Item/Service")})
                    entries.append({"account": "Output GST", "debit": 0, "credit": gst_amt, "description": "GST"})
                
                # Update inventory if Item contains HSN
                if row.get("HSN Code"):
                    item = await db.inventory.find_one({"hsn_code": row.get("HSN Code")}, {"_id": 0})
                    if not item:
                        # Create new inventory item
                        new_item = InventoryItem(
                            item_name=row.get("Item/Service"),
                            hsn_code=row.get("HSN Code"),
                            landed_cost=base,
                            current_stock=float(row.get("Quantity", 1))
                        )
                        await db.inventory.insert_one(new_item.model_dump())
                    else:
                        # Update stock
                        qty = float(row.get("Quantity", 1))
                        await db.inventory.update_one(
                            {"hsn_code": row.get("HSN Code")},
                            {"$inc": {"current_stock": qty}}
                        )
            
            txn_dict = txn.model_dump()
            txn_dict["journal_entries"] = entries
            await db.transactions.insert_one(txn_dict)
            imported += 1
        
        return {"success": True, "imported": imported, "validation": validation}
    except Exception as e:
        logging.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Document Upload
@api_router.post("/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    """Upload invoice/receipt and perform OCR"""
    try:
        ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
        doc_id = str(uuid.uuid4())
        path = f"{APP_NAME}/documents/{doc_id}.{ext}"
        
        data = await file.read()
        result = put_object(path, data, file.content_type or "application/octet-stream")
        
        extracted_data = await extract_document_data(data, file.filename)
        
        doc_record = DocumentUpload(
            id=doc_id,
            storage_path=result["path"],
            original_filename=file.filename,
            content_type=file.content_type,
            size=result["size"],
            extracted_data=extracted_data
        )
        
        await db.documents.insert_one(doc_record.model_dump())
        
        return {"document_id": doc_id, "extracted_data": extracted_data}
    except Exception as e:
        logging.error(f"Document upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documents/{doc_id}/download")
async def download_document(doc_id: str):
    """Download document file"""
    try:
        record = await db.documents.find_one({"id": doc_id, "is_deleted": False}, {"_id": 0})
        if not record:
            raise HTTPException(status_code=404, detail="Document not found")
        
        data, content_type = get_object(record["storage_path"])
        return Response(content=data, media_type=record.get("content_type", content_type))
    except Exception as e:
        logging.error(f"Document download error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Process Prompt
@api_router.post("/transactions/prompt")
async def process_prompt(request: PromptRequest):
    """Process natural language prompt and create draft transaction"""
    try:
        extracted_ocr = None
        if request.document_id:
            doc = await db.documents.find_one({"id": request.document_id}, {"_id": 0})
            if doc:
                extracted_ocr = doc.get("extracted_data")
        
        parsed_data = await parse_prompt_with_ai(request.prompt, request.module, extracted_ocr, request.cost_center)
        
        transaction = TransactionBase(
            user_prompt=request.prompt,
            user_id=request.user_id,
            module=request.module,
            status="draft",
            posting_date=parsed_data.get("posting_date"),
            service_period_start=parsed_data.get("service_period_start"),
            service_period_end=parsed_data.get("service_period_end"),
            business_unit=parsed_data.get("business_unit"),
            cost_center=request.cost_center,
            extracted_data=parsed_data,
            journal_entries=parsed_data.get("journal_entries", []),
            document_id=request.document_id
        )
        
        await db.transactions.insert_one(transaction.model_dump())
        
        return transaction.model_dump()
    except Exception as e:
        logging.error(f"Prompt processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/transactions/drafts")
async def get_drafts(module: Optional[str] = None):
    """Get all draft transactions"""
    query = {"status": "draft"}
    if module:
        query["module"] = module
    
    drafts = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return drafts

@api_router.post("/transactions/post")
async def post_transaction(request: PostTransactionRequest):
    """Post a draft transaction to ledger"""
    try:
        transaction = await db.transactions.find_one({"id": request.transaction_id}, {"_id": 0})
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        if transaction["status"] != "draft":
            raise HTTPException(status_code=400, detail="Transaction already posted")
        
        cost_center = request.cost_center or transaction.get("cost_center", "General")
        
        await db.transactions.update_one(
            {"id": request.transaction_id},
            {"$set": {"status": "posted", "posted_at": datetime.now(timezone.utc).isoformat(), "cost_center": cost_center}}
        )
        
        journal_entries = transaction.get("journal_entries", [])
        for entry in journal_entries:
            je = JournalEntry(
                transaction_id=request.transaction_id,
                account=entry.get("account", ""),
                debit=entry.get("debit", 0.0),
                credit=entry.get("credit", 0.0),
                description=entry.get("description", ""),
                posting_date=transaction["posting_date"],
                cost_center=cost_center,
                tax_details=entry.get("tax_details")
            )
            await db.journal_entries.insert_one(je.model_dump())
            
            # Update CoA balances
            account_name = entry.get("account")
            debit_amt = entry.get("debit", 0.0)
            credit_amt = entry.get("credit", 0.0)
            net_change = debit_amt - credit_amt
            
            await db.chart_of_accounts.update_one(
                {"ledger_name": account_name},
                {"$inc": {"current_balance": net_change}},
                upsert=False
            )
        
        return {"message": "Transaction posted successfully", "transaction_id": request.transaction_id}
    except Exception as e:
        logging.error(f"Post transaction error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/transactions/posted")
async def get_posted_transactions(module: Optional[str] = None, limit: int = 50):
    """Get posted transactions"""
    query = {"status": "posted"}
    if module:
        query["module"] = module
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("posted_at", -1).to_list(limit)
    return transactions

# Conversational Reporting
@api_router.post("/reports/query")
async def conversational_report(request: ReportQuery):
    """AI-powered conversational reporting"""
    try:
        query = {"status": "posted"}
        if request.start_date and request.end_date:
            query["posting_date"] = {"$gte": request.start_date, "$lte": request.end_date}
        
        transactions = await db.transactions.find(query, {"_id": 0}).to_list(1000)
        journal_entries = await db.journal_entries.find({}, {"_id": 0}).to_list(1000)
        
        chat = LlmChat(
            api_key=EMERGENT_KEY,
            session_id=f"report-{uuid.uuid4()}",
            system_message=f"""You are an AI financial analyst for Kairos Accounting system (India ERP). 
            Answer the user's query based on the provided transaction data.
            
            Transactions: {json.dumps(transactions[:100])}
            Journal Entries: {json.dumps(journal_entries[:100])}
            
            Provide clear, concise answers with numbers and insights."""
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")
        
        user_message = UserMessage(text=request.query)
        response = await chat.send_message(user_message)
        
        return {"query": request.query, "answer": response, "data_points": len(transactions)}
    except Exception as e:
        logging.error(f"Report query error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Financial Statements
@api_router.get("/reports/balance-sheet")
async def get_balance_sheet(as_of_date: Optional[str] = None, cost_center: Optional[str] = None):
    """Generate Balance Sheet"""
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).date().isoformat()
    
    query = {"posting_date": {"$lte": as_of_date}}
    if cost_center:
        query["cost_center"] = cost_center
    
    entries = await db.journal_entries.find(query, {"_id": 0}).to_list(10000)
    
    balances = {}
    for entry in entries:
        account = entry["account"]
        if account not in balances:
            balances[account] = {"debit": 0, "credit": 0}
        balances[account]["debit"] += entry["debit"]
        balances[account]["credit"] += entry["credit"]
    
    assets = {k: v for k, v in balances.items() if "Asset" in k or "Bank" in k or "Cash" in k}
    liabilities = {k: v for k, v in balances.items() if "Liability" in k or "Payable" in k}
    equity = {k: v for k, v in balances.items() if "Equity" in k or "Capital" in k}
    
    # TB Integrity Check
    total_debit = sum(b["debit"] for b in balances.values())
    total_credit = sum(b["credit"] for b in balances.values())
    difference = abs(total_debit - total_credit)
    has_suspense = difference > 0.01
    
    return {
        "as_of_date": as_of_date,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "total_entries": len(entries),
        "suspense_account": has_suspense,
        "suspense_amount": difference if has_suspense else 0,
        "cost_center": cost_center
    }

@api_router.get("/reports/profit-loss")
async def get_profit_loss(start_date: str, end_date: str, cost_center: Optional[str] = None):
    """Generate Profit & Loss Statement"""
    query = {"posting_date": {"$gte": start_date, "$lte": end_date}}
    if cost_center:
        query["cost_center"] = cost_center
    
    entries = await db.journal_entries.find(query, {"_id": 0}).to_list(10000)
    
    revenue = {}
    expenses = {}
    
    for entry in entries:
        account = entry["account"]
        if "Revenue" in account or "Income" in account or "Sales" in account:
            if account not in revenue:
                revenue[account] = 0
            revenue[account] += entry["credit"] - entry["debit"]
        elif "Expense" in account or "Cost" in account or "Purchases" in account:
            if account not in expenses:
                expenses[account] = 0
            expenses[account] += entry["debit"] - entry["credit"]
    
    total_revenue = sum(revenue.values())
    total_expenses = sum(expenses.values())
    
    return {
        "period": {"start": start_date, "end": end_date},
        "revenue": revenue,
        "total_revenue": total_revenue,
        "expenses": expenses,
        "total_expenses": total_expenses,
        "net_profit": total_revenue - total_expenses,
        "cost_center": cost_center
    }

@api_router.get("/reports/trial-balance")
async def get_trial_balance(as_of_date: Optional[str] = None):
    """Generate Trial Balance"""
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).date().isoformat()
    
    entries = await db.journal_entries.find(
        {"posting_date": {"$lte": as_of_date}},
        {"_id": 0}
    ).to_list(10000)
    
    balances = {}
    for entry in entries:
        account = entry["account"]
        if account not in balances:
            balances[account] = {"debit": 0, "credit": 0}
        balances[account]["debit"] += entry["debit"]
        balances[account]["credit"] += entry["credit"]
    
    total_debit = sum(b["debit"] for b in balances.values())
    total_credit = sum(b["credit"] for b in balances.values())
    
    return {
        "as_of_date": as_of_date,
        "accounts": balances,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "difference": total_debit - total_credit,
        "in_balance": abs(total_debit - total_credit) < 0.01
    }

# PDF Export
@api_router.get("/reports/balance-sheet/pdf")
async def download_balance_sheet_pdf(as_of_date: Optional[str] = None):
    """Download Balance Sheet as PDF"""
    data = await get_balance_sheet(as_of_date)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"<b>Balance Sheet - Kairos Accounting</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Date
    date_para = Paragraph(f"As of: {data['as_of_date']}", styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 12))
    
    # Assets Table
    if data['assets']:
        assets_data = [["Assets", "Amount"]]
        for acc, bal in data['assets'].items():
            assets_data.append([acc, f"₹{bal['debit'] - bal['credit']:.2f}"])
        
        t = Table(assets_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(content=buffer.read(), media_type="application/pdf", 
                   headers={"Content-Disposition": f"attachment; filename=balance_sheet_{as_of_date}.pdf"})

# Excel Export
@api_router.get("/reports/trial-balance/excel")
async def download_trial_balance_excel(as_of_date: Optional[str] = None):
    """Download Trial Balance as Excel"""
    data = await get_trial_balance(as_of_date)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    # Header
    ws['A1'] = "Kairos Accounting - Trial Balance"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"As of: {data['as_of_date']}"
    
    # Column headers
    ws['A4'] = "Account"
    ws['B4'] = "Debit"
    ws['C4'] = "Credit"
    
    for cell in ['A4', 'B4', 'C4']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    row = 5
    for acc, bal in data['accounts'].items():
        ws[f'A{row}'] = acc
        ws[f'B{row}'] = bal['debit']
        ws[f'C{row}'] = bal['credit']
        row += 1
    
    # Totals
    ws[f'A{row}'] = "Total"
    ws[f'B{row}'] = data['total_debit']
    ws[f'C{row}'] = data['total_credit']
    ws[f'A{row}'].font = Font(bold=True)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(content=buffer.read(), 
                   media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": f"attachment; filename=trial_balance_{as_of_date}.xlsx"})

# ==================== INTEGRATE NEW ERP MODULES ====================
# Import route modules
try:
    from routes_crm import router as crm_router, set_db as crm_set_db, set_ai_orchestrator as crm_set_ai
    from routes_sales import router as sales_router, set_db as sales_set_db, set_ai_orchestrator as sales_set_ai
    from routes_stock import router as stock_router, set_db as stock_set_db, set_ai_orchestrator as stock_set_ai
    from routes_hr import router as hr_router, set_db as hr_set_db
    from routes_purchase import router as purchase_router, set_db as purchase_set_db
    from routes_selling import router as selling_router, set_db as selling_set_db
    from routes_financial_statements import router as fs_router, set_db as fs_set_db
    from routes_statutory import router as stat_router, set_db as stat_set_db
    from routes_manufacturing import router as mfg_router, set_db as mfg_set_db
    from routes_company import router as company_router, set_db as company_set_db, set_key as company_set_key
    from routes_audit import router as audit_router, set_db as audit_set_db
    from routes_gst import router as gst_router, set_key as gst_set_key
    from routes_aging import router as aging_router, set_db as aging_set_db
    import audit_trail
    from ai_orchestrator import AIOrchestrator
    
    # Initialize AI Orchestrator
    ai_orch = AIOrchestrator(EMERGENT_KEY)
    
    # Set database and AI orchestrator for all modules
    crm_set_db(db)
    crm_set_ai(ai_orch)
    sales_set_db(db)
    sales_set_ai(ai_orch)
    stock_set_db(db)
    stock_set_ai(ai_orch)
    hr_set_db(db)
    purchase_set_db(db)
    selling_set_db(db)
    fs_set_db(db)
    stat_set_db(db)
    mfg_set_db(db)
    company_set_db(db)
    company_set_key(EMERGENT_KEY)
    audit_set_db(db)
    audit_trail.set_db(db)
    gst_set_key(EMERGENT_KEY)
    aging_set_db(db)
    
    # Universal AI Prompt Endpoint
    @api_router.post("/ai/universal-prompt")
    async def universal_ai_prompt(data: dict):
        """Universal AI prompt that routes to appropriate module"""
        try:
            prompt = data.get("prompt")
            context = data.get("context", {})
            
            result = await ai_orch.process_universal_prompt(prompt, context)
            return result
        except Exception as e:
            logging.error(f"Universal prompt error: {e}")
            raise HTTPException(status_code=500, detail=str(e))

    # ═══════════════════════════════════════════════════
    # AI PARSE PROMPT — the core of the AI-first entry
    # ═══════════════════════════════════════════════════
    @api_router.post("/ai/parse-prompt")
    async def ai_parse_prompt(body: dict):
        """Parse a natural language prompt into structured ERP form data"""
        prompt = body.get("prompt", "").strip()
        if not prompt:
            raise HTTPException(status_code=400, detail="Prompt is required")

        # 1. Fetch master data for context
        vendors_raw = await db.entities.find({"entity_type": "vendor"}, {"_id": 0}).to_list(200)
        customers_raw = await db.entities.find({"entity_type": "customer"}, {"_id": 0}).to_list(200)
        items_raw = await db.items.find({}, {"_id": 0}).to_list(200)
        cost_centers_raw = await db.cost_centers.find({}, {"_id": 0}).to_list(50)
        coa_raw = await db.chart_of_accounts.find({"is_active": {"$ne": False}}, {"_id": 0, "ledger_name": 1, "category": 1}).to_list(500)
        pending_pos = await db.purchase_orders.find(
            {"grn_status": "Pending"}, {"_id": 0, "id": 1, "po_number": 1, "vendor": 1, "items": 1, "grand_total": 1}
        ).to_list(100)
        pending_sos = await db.selling_sales_orders.find(
            {"delivery_status": {"$ne": "Fully Delivered"}, "status": {"$ne": "Cancelled"}},
            {"_id": 0, "id": 1, "so_number": 1, "customer": 1, "items": 1, "grand_total": 1}
        ).to_list(100)

        vendor_names = [v.get("name", "") for v in vendors_raw]
        customer_names = [c.get("name", "") for c in customers_raw]
        item_summaries = [{"code": i.get("item_code",""), "name": i.get("item_name",""), "uom": i.get("uom","KG"), "rate": i.get("valuation_rate",0), "hsn_sac": i.get("hsn_sac",""), "gst_rate": i.get("gst_rate",18)} for i in items_raw]
        vendor_summaries = [{"name": v.get("name",""), "gstin": v.get("gstin",""), "state": v.get("state",""), "state_code": v.get("gst_state_code","")} for v in vendors_raw]
        customer_summaries = [{"name": c.get("name",""), "gstin": c.get("gstin",""), "state": c.get("state",""), "state_code": c.get("gst_state_code","")} for c in customers_raw]
        cc_names = [c.get("name","") for c in cost_centers_raw]
        ledger_names = [c.get("ledger_name","") for c in coa_raw]
        po_summaries = [{"id": p["id"], "number": p.get("po_number",""), "vendor": p.get("vendor",""), "total": p.get("grand_total",0)} for p in pending_pos]
        so_summaries = [{"id": s["id"], "number": s.get("so_number",""), "customer": s.get("customer",""), "total": s.get("grand_total",0)} for s in pending_sos]

        system_msg = f"""You are the AI brain of Kairos AI ERP for PolyMerx Specialty Chemicals Pvt. Ltd.

MASTER DATA (use these exact names when matching):
- Vendors: {json.dumps(vendor_summaries)}
- Customers: {json.dumps(customer_summaries)}
- Items: {json.dumps(item_summaries)}
- Cost Centers: {json.dumps(cc_names)}
- Ledger Accounts: {json.dumps(ledger_names[:60])}
- Pending POs (for GRN): {json.dumps(po_summaries)}
- Pending SOs (for Delivery): {json.dumps(so_summaries)}

Parse the user's prompt and return a JSON object with:
{{
  "intent": "<one of: purchase_order, sales_order, work_order, journal_entry, goods_receipt, delivery_note, purchase_invoice, sales_invoice, vendor_payment, customer_receipt, crm_lead>",
  "confidence": 0.0 to 1.0,
  "summary": "One-line summary of what will be created",
  "extracted": {{ <fields extracted from the prompt, matching master data names exactly> }},
  "missing": ["<list of required field keys that were NOT in the prompt>"]
}}

FIELD SCHEMAS by intent:

purchase_order:
  required: vendor, items (array of {{item_code, item_name, qty, rate, uom, amount, hsn_sac, gst_rate}}), cost_center
  optional: delivery_date, payment_terms
  auto: vendor_state and company_state used for CGST+SGST vs IGST determination

sales_order:
  required: customer, items (array of {{item_code, item_name, qty, rate, uom, amount, hsn_sac, gst_rate}}), cost_center
  optional: delivery_date, payment_terms, po_no
  auto: customer_state and company_state used for CGST+SGST vs IGST determination

work_order:
  required: production_item (item_code of FG), qty_to_produce, cost_center
  optional: bom_items (array of {{item_code, qty, rate}}), planned_start, planned_end

journal_entry:
  required: entries (array of {{account, debit, credit, description}}), narration
  optional: posting_date, cost_center

goods_receipt:
  required: po_id (from pending POs list)
  optional: received_qty (per item, defaults to PO qty)

delivery_note:
  required: so_id (from pending SOs list)
  optional: delivered_qty (per item, defaults to SO qty)

crm_lead:
  required: company, contact_name
  optional: phone, email, interest, source, est_value

RULES:
- Fuzzy match entity/item names to the closest master data entry. E.g. "Aditya Birla" → match to vendor list.
- If an item name like "Epoxy Resin" is mentioned, find the closest item_code from master data.
- For items array, always compute amount = qty * rate.
- Default cost_center to the most relevant one. Default gst_rate to 18.
- Use today's date ({datetime.now(timezone.utc).date().isoformat()}) if dates aren't specified.
- For goods_receipt/delivery_note, match PO/SO by number or vendor/customer name.
- Return ONLY valid JSON. No markdown fences."""

        try:
            chat = LlmChat(
                api_key=EMERGENT_KEY,
                session_id=f"parse-{uuid.uuid4()}",
                system_message=system_msg
            ).with_model("anthropic", "claude-sonnet-4-5-20250929")

            raw = await chat.send_message(UserMessage(text=prompt))
            from ai_orchestrator import clean_json_response
            parsed = clean_json_response(raw)

            # Enrich response with master data lists for frontend dropdowns
            parsed["master_data"] = {
                "vendors": vendor_names,
                "customers": customer_names,
                "items": item_summaries,
                "cost_centers": cc_names,
                "ledgers": ledger_names,
                "pending_pos": po_summaries,
                "pending_sos": so_summaries,
            }
            return parsed
        except Exception as e:
            logging.error(f"AI parse-prompt error: {e}")
            raise HTTPException(status_code=500, detail=f"AI parsing failed: {str(e)}")
    
    # Include routers
    api_router.include_router(crm_router)
    api_router.include_router(sales_router)
    api_router.include_router(stock_router)
    api_router.include_router(hr_router)
    api_router.include_router(purchase_router)
    api_router.include_router(selling_router)
    api_router.include_router(fs_router)
    api_router.include_router(stat_router)
    api_router.include_router(mfg_router)
    api_router.include_router(company_router, prefix="/company")
    api_router.include_router(audit_router)
    api_router.include_router(gst_router)
    api_router.include_router(aging_router)
    
    # New modules: Projects, Timesheets, Revenue Recognition
    from routes_projects import router as projects_router
    from routes_timesheets import router as timesheets_router
    from routes_revenue import router as revenue_router
    api_router.include_router(projects_router)
    api_router.include_router(timesheets_router)
    api_router.include_router(revenue_router)
    
    # AI Agents
    from routes_agents import router as agents_router, set_config as set_agents_config
    set_agents_config(os.environ.get("EMERGENT_LLM_KEY"), db)
    api_router.include_router(agents_router)
    
    # Bank Reconciliation
    from routes_bank_recon import router as bank_recon_router, set_db as set_bank_recon_db
    set_bank_recon_db(db)
    api_router.include_router(bank_recon_router)
    
    # Employee Analytics
    from routes_employee_analytics import router as emp_analytics_router, set_db as set_emp_analytics_db
    set_emp_analytics_db(db)
    api_router.include_router(emp_analytics_router)
    
    
    # Leave Management
    from routes_leave_management import router as leave_management_router, set_db as set_leave_management_db
    set_leave_management_db(db)
    api_router.include_router(leave_management_router)
    
    # Expense Management
    from routes_expense_management import router as expense_management_router, set_db as set_expense_management_db
    set_expense_management_db(db)
    api_router.include_router(expense_management_router)
    
    # Feedback
    from routes_feedback import router as feedback_router, set_db as set_feedback_db
    set_feedback_db(db)
    api_router.include_router(feedback_router)
    
    # Announcements
    from routes_announcements import router as announcements_router, set_db as set_announcements_db
    set_announcements_db(db)
    api_router.include_router(announcements_router)
    
    # Chart Of Accounts
    from routes_chart_of_accounts import router as chart_of_accounts_router, set_db as set_chart_of_accounts_db
    set_chart_of_accounts_db(db)
    api_router.include_router(chart_of_accounts_router)
    
    # Vendors
    from routes_vendors import router as vendors_router, set_db as set_vendors_db
    set_vendors_db(db)
    api_router.include_router(vendors_router)
    
    # Customers
    from routes_customers import router as customers_router, set_db as set_customers_db
    set_customers_db(db)
    api_router.include_router(customers_router)
    # ── Advanced Enterprise Modules ──
    from routes_approvals import router as approvals_router, set_db as set_approvals_db
    set_approvals_db(db)
    api_router.include_router(approvals_router)

    from routes_budgets import router as budgets_router, set_db as set_budgets_db
    set_budgets_db(db)
    api_router.include_router(budgets_router)

    from routes_contracts import router as contracts_router, set_db as set_contracts_db
    set_contracts_db(db)
    api_router.include_router(contracts_router)

    from routes_resources import router as resources_router, set_db as set_resources_db
    set_resources_db(db)
    api_router.include_router(resources_router)

    from routes_forex import router as forex_router, set_db as set_forex_db
    set_forex_db(db)
    api_router.include_router(forex_router)

    from routes_billing import router as billing_router, set_db as set_billing_db
    set_billing_db(db)
    api_router.include_router(billing_router)

    from routes_documents import router as docs_router, set_db as set_docs_db
    set_docs_db(db)
    api_router.include_router(docs_router)

    from routes_notifications import router as notifications_router, set_db as set_notifications_db
    set_notifications_db(db)
    api_router.include_router(notifications_router)

    from routes_compliance import router as compliance_router, set_db as set_compliance_db
    set_compliance_db(db)
    api_router.include_router(compliance_router)

    from routes_portal import router as portal_router, set_db as set_portal_db
    set_portal_db(db)
    api_router.include_router(portal_router)

    logging.info("ERP modules integrated (including 10 advanced modules)")
except Exception as e:
    logging.error(f"Failed to integrate ERP modules: {e}")

# ==================== MANUAL JOURNAL ENTRIES ====================
@api_router.post("/journal-entries/manual")
async def create_manual_journal_entry(data: dict):
    """Create manual journal entry for corrections, adjustments, or audit entries"""
    entry = {
        "id": str(uuid.uuid4()),
        "entry_type": data.get("entry_type", "Manual Entry"),
        "posting_date": data.get("posting_date", datetime.now(timezone.utc).date().isoformat()),
        "reference_date": data.get("reference_date"),
        "cost_center": data.get("cost_center", "General"),
        "journal_entries": data.get("journal_entries", []),
        "narration": data.get("narration", ""),
        "reference_transaction_id": data.get("reference_transaction_id"),
        "voucher_type": data.get("voucher_type", "Journal Entry"),
        "status": "Draft",
        "user_id": data.get("user_id", "default_user"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    total_debit = sum(je.get("debit", 0) for je in entry["journal_entries"])
    total_credit = sum(je.get("credit", 0) for je in entry["journal_entries"])
    
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(status_code=400, detail=f"Not balanced. Debit: {total_debit}, Credit: {total_credit}")
    
    await db.manual_journal_entries.insert_one(entry)
    del entry["_id"]
    await audit_trail.log_audit(audit_trail.ACTION_CREATE, audit_trail.DOC_MANUAL_JE, entry["id"], entry["id"][:12], snapshot=entry, notes=f"Manual JE: {entry.get('narration', '')[:80]}")
    return entry

@api_router.get("/journal-entries/manual")
async def get_manual_journal_entries(entry_type: Optional[str] = None, status: Optional[str] = None, limit: int = 100):
    query = {}
    if entry_type:
        query["entry_type"] = entry_type
    if status:
        query["status"] = status
    entries = await db.manual_journal_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return entries

@api_router.post("/journal-entries/manual/{entry_id}/post")
async def post_manual_journal_entry(entry_id: str):
    entry = await db.manual_journal_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry["status"] != "Draft":
        raise HTTPException(status_code=400, detail="Already posted")
    
    for je in entry["journal_entries"]:
        journal_entry = JournalEntry(
            transaction_id=entry_id,
            account=je.get("account"),
            debit=je.get("debit", 0.0),
            credit=je.get("credit", 0.0),
            description=je.get("description", entry.get("narration", "")),
            posting_date=entry["posting_date"],
            cost_center=entry["cost_center"]
        )
        await db.journal_entries.insert_one(journal_entry.model_dump())
        net_change = je.get("debit", 0.0) - je.get("credit", 0.0)
        await db.chart_of_accounts.update_one(
            {"ledger_name": je.get("account")},
            {"$inc": {"current_balance": net_change}},
            upsert=False
        )
    
    await db.manual_journal_entries.update_one(
        {"id": entry_id},
        {"$set": {"status": "Posted", "posted_at": datetime.now(timezone.utc).isoformat()}}
    )
    await audit_trail.log_audit(audit_trail.ACTION_POST, audit_trail.DOC_MANUAL_JE, entry_id, entry_id[:12], changes=[{"field": "status", "old_value": "Draft", "new_value": "Posted"}], notes=f"Manual JE posted: {entry.get('narration', '')[:80]}")
    return {"message": "Posted successfully"}

# ==================== ADMIN DATA TABLES ====================
@api_router.get("/admin/tables")
async def get_all_tables():
    collections = await db.list_collection_names()
    tables = [c for c in collections if not c.startswith('system.')]
    table_info = []
    for table in tables:
        count = await db[table].count_documents({})
        table_info.append({"name": table, "count": count})
    return sorted(table_info, key=lambda x: x['name'])

@api_router.get("/admin/tables/{table_name}")
async def get_table_data(table_name: str, skip: int = 0, limit: int = 100, search: Optional[str] = None):
    try:
        query = {}
        if search:
            query = {
                "$or": [
                    {"id": {"$regex": search, "$options": "i"}},
                    {"name": {"$regex": search, "$options": "i"}},
                    {"customer_name": {"$regex": search, "$options": "i"}},
                    {"employee_name": {"$regex": search, "$options": "i"}},
                ]
            }
        total = await db[table_name].count_documents(query)
        records = await db[table_name].find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
        return {"table": table_name, "total": total, "skip": skip, "limit": limit, "records": records}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/admin/tables/{table_name}/export")
async def export_table_data(table_name: str):
    try:
        records = await db[table_name].find({}, {"_id": 0}).to_list(10000)
        if not records:
            return {"message": "No data"}
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={table_name}.csv"}
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Include router
routes_auth.set_db(db)
api_router.include_router(routes_auth.router)
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Kairos Accounting - Storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
    try:
        await routes_auth.seed_users()
    except Exception as e:
        logger.error(f"User seeding failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
